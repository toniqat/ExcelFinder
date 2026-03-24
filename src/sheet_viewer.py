import os
import pandas as pd
import warnings
import tempfile
import shutil
import subprocess
import stat
import platform
import time
from pathlib import Path
from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel, QCheckBox,
                            QTableWidget, QTableWidgetItem, QHeaderView,
                            QMessageBox, QSizePolicy, QPushButton)
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QBrush, QColor

from constants import FILE_OPEN_TIMEOUT

# 프로세스 모니터링을 위한 라이브러리 (선택적 import)
try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False

# Windows 전용 라이브러리 (선택적 import)
if platform.system() == "Windows":
    try:
        import win32gui
        import win32process
        WIN32_AVAILABLE = True
    except ImportError:
        WIN32_AVAILABLE = False
else:
    WIN32_AVAILABLE = False

class SheetViewer(QDialog):
    """엑셀 시트에서 검색된 행을 보여주는 다이얼로그"""
    def __init__(self, file_path, sheet_name, row_index, parent=None, cached_data=None,
                 highlight_col=None, search_keyword=''):
        super().__init__(parent)
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.row_index = row_index  # 엑셀에서의 행 인덱스 (1-based)
        self.hide_nan = True  # 기본적으로 NaN 값을 숨김
        self.df = None
        self.header_row = None
        self.cached_data = cached_data  # 캐시된 데이터 (header_data, row_data)
        self.highlight_col = highlight_col  # 하이라이트할 열 인덱스 (0-based)
        self.search_keyword = search_keyword  # 검색 키워드

        # Help 버튼 제거
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)

        self.init_ui()
        
        # 캐시된 데이터가 있으면 사용, 없으면 파일에서 로드
        if self.cached_data:
            self.load_from_cache()
        else:
            self.load_row_data()
        
    def init_ui(self):
        # 다이얼로그 설정
        self.setWindowTitle(f'{os.path.basename(self.file_path)} > {self.sheet_name}')
        self.setGeometry(150, 150, 900, 300)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        
        # 레이아웃 설정
        layout = QVBoxLayout(self)
        
        # 정보 레이블
        info_label = QLabel(f"파일: {os.path.basename(self.file_path)}, 시트: {self.sheet_name}, 행: {self.row_index}")
        layout.addWidget(info_label)
        
        # 파일 작업 버튼들
        button_layout = QHBoxLayout()
        
        # 파일 열기 버튼
        self.open_file_btn = QPushButton("파일 열기")
        self.open_file_btn.setToolTip("기본 프로그램으로 파일 열기")
        self.open_file_btn.clicked.connect(self.open_file)
        button_layout.addWidget(self.open_file_btn)
        
        # 읽기 전용 해제 후 열기 버튼
        self.open_file_writable_btn = QPushButton("읽기 전용 해제 후 열기")
        self.open_file_writable_btn.setToolTip("파일의 읽기 전용 속성을 해제한 후 열기")
        self.open_file_writable_btn.clicked.connect(self.open_file_writable)
        button_layout.addWidget(self.open_file_writable_btn)
        
        # 폴더 열기 버튼
        self.open_folder_btn = QPushButton("폴더 탐색기에서 열기")
        self.open_folder_btn.setToolTip("파일이 있는 폴더를 탐색기로 열기")
        self.open_folder_btn.clicked.connect(self.open_folder)
        button_layout.addWidget(self.open_folder_btn)
        
        # 버튼 레이아웃을 메인 레이아웃에 추가
        layout.addLayout(button_layout)
        
        # NaN 값 숨김 체크박스
        self.hide_nan_checkbox = QCheckBox("빈 값(NaN) 숨기기")
        self.hide_nan_checkbox.setChecked(True)
        self.hide_nan_checkbox.stateChanged.connect(self.toggle_hide_nan)
        layout.addWidget(self.hide_nan_checkbox)
        
        # 테이블 위젯 생성
        self.table = QTableWidget()
        # 행 번호(vertical header) 숨기기
        self.table.verticalHeader().setVisible(False)
        layout.addWidget(self.table)
        
        # 창 크기 조절 시 테이블 크기도 함께 조절
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
    def resizeEvent(self, event):
        # 창 크기가 조절될 때 테이블 크기도 조절
        super().resizeEvent(event)
        
    def toggle_hide_nan(self, state):
        """NaN 값 숨김 토글"""
        self.hide_nan = (state == Qt.Checked)
        self.update_table()
        
    def update_table(self):
        """테이블 업데이트"""
        if self.df is None or self.header_row is None:
            return
            
        # 행 데이터 가져오기 - 인덱스 조정 (1-based를 0-based로 변환)
        row_data = self.df.iloc[self.row_index - 1]
        
        # 표시할 열 결정
        visible_columns = []
        for col_idx, value in enumerate(row_data):
            if not self.hide_nan or not (pd.isna(value) or value == '' or value is None):
                visible_columns.append(col_idx)
                
        # 테이블 설정
        self.table.clear()
        self.table.setRowCount(1)
        self.table.setColumnCount(len(visible_columns))
        
        # 헤더 설정
        header_labels = []
        for i, col_idx in enumerate(visible_columns):
            header_labels.append(str(self.header_row[col_idx]))
        self.table.setHorizontalHeaderLabels(header_labels)
        
        # 데이터 행 추가
        for i, col_idx in enumerate(visible_columns):
            value = row_data[col_idx]
            item = QTableWidgetItem(str(value) if not pd.isna(value) else "")

            # 검색된 열에 하이라이트 적용
            if self.highlight_col is not None and col_idx == self.highlight_col:
                item.setBackground(QBrush(QColor(255, 255, 0, 100)))  # 노란색 배경

            self.table.setItem(0, i, item)
            
        # 열 너비 조정 - 사용자가 조절할 수 있도록 설정
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        
    def load_from_cache(self):
        """캐시된 데이터로부터 행 데이터 로드"""
        try:
            if self.cached_data:
                header_data, row_data = self.cached_data
                
                # 캐시된 데이터로 DataFrame 생성
                # 헤더 행과 데이터 행만 포함하는 간단한 DataFrame 생성
                data = [header_data, row_data]
                self.df = pd.DataFrame(data)
                
                # 첫 번째 행은 헤더로 사용
                self.header_row = self.df.iloc[0]
                
                # 캐시된 데이터는 항상 2행만 있음 (헤더 + 데이터)
                # row_index는 무시하고 항상 두 번째 행(인덱스 1)을 사용
                self.row_index = 2  # 1-based 인덱스로 두 번째 행은 2
                
                # 테이블 업데이트
                self.update_table()
                
        except Exception as e:
            QMessageBox.warning(self, '경고', f'캐시된 데이터 로딩 중 오류 발생: {str(e)}\n파일에서 직접 로드합니다.')
            # 캐시 로드 실패 시 파일에서 직접 로드
            self.load_row_data()
    
    def load_row_data(self):
        """파일에서 직접 행 데이터 로드"""
        try:
            # 경고 무시 설정
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                
                # 파일 확장자 확인
                file_ext = Path(self.file_path).suffix.lower()
                
                # CSV/TSV 파일 처리
                if file_ext in ['.csv', '.tsv']:
                    try:
                        # CSV 파일을 스마트하게 읽기 (file_processor의 함수 사용)
                        from file_processor import read_csv_smart
                        self.df = read_csv_smart(self.file_path)
                        
                        if self.df.empty:
                            QMessageBox.warning(self, '경고', 'CSV 파일이 비어있거나 읽을 수 없습니다.')
                            return
                        
                        # CSV는 첫 번째 행이 헤더
                        self.header_row = self.df.columns
                        
                        # 테이블 업데이트
                        self.update_table()
                        return
                        
                    except Exception as e:
                        QMessageBox.critical(self, '오류', f'CSV 파일 로딩 중 오류 발생: {str(e)}')
                        return
                
                # 시트 데이터 로드 (수식 대신 값만 참조, 조건부 서식 무시)
                # 파일 확장자에 따라 적절한 엔진 선택
                try:
                    if file_ext == '.xls':
                        # .xls 파일은 xlrd 엔진 사용
                        self.df = pd.read_excel(self.file_path, sheet_name=self.sheet_name, header=None,
                                              engine='xlrd')
                    elif file_ext == '.xlsb':
                        # .xlsb 파일은 pyxlsb 엔진 사용 (필요시 설치)
                        try:
                            import pyxlsb
                            self.df = pd.read_excel(self.file_path, sheet_name=self.sheet_name, header=None,
                                                  engine='pyxlsb')
                        except ImportError:
                            # pyxlsb가 설치되지 않은 경우 openpyxl 시도
                            self.df = pd.read_excel(self.file_path, sheet_name=self.sheet_name, header=None,
                                                  engine='openpyxl')
                    else:
                        # .xlsx, .xlsm 파일은 openpyxl 엔진 사용
                        # pandas의 read_excel은 data_only 파라미터를 직접 지원하지 않음
                        from openpyxl import load_workbook
                        wb = load_workbook(self.file_path, data_only=True, read_only=True)
                        self.df = pd.read_excel(self.file_path, sheet_name=self.sheet_name, header=None,
                                              engine='openpyxl')
                except Exception as e1:
                    # 두 번째 방법: 다양한 옵션 시도
                    try:
                        if file_ext == '.xls':
                                # xlrd 설정 변경
                                import xlrd
                                xlrd.mmap_mode = False  # 메모리 매핑 모드 비활성화
                                
                                # Excel 97-2003 워크시트(.xls) 파일 처리를 위한 특별 설정
                                try:
                                    # 첫 번째 시도: 기본 설정으로 열기
                                    workbook = xlrd.open_workbook(self.file_path, formatting_info=False, 
                                                                 on_demand=True)
                                except Exception as xls_e1:
                                    try:
                                        # 두 번째 시도: 인코딩 지정 (한글 파일)
                                        workbook = xlrd.open_workbook(self.file_path, formatting_info=False, 
                                                                     encoding_override='cp949',
                                                                     on_demand=True)
                                    except Exception as xls_e2:
                                        try:
                                            # 세 번째 시도: 레거시 모드 (Excel 97-2003 전용)
                                            # ragged_rows=True: 행마다 열 수가 다를 수 있음을 허용
                                            workbook = xlrd.open_workbook(self.file_path, formatting_info=False, 
                                                                         ragged_rows=True,
                                                                         on_demand=True)
                                        except Exception as xls_e3:
                                            # 네 번째 시도: 모든 옵션 조합
                                            workbook = xlrd.open_workbook(self.file_path, formatting_info=False, 
                                                                         encoding_override='cp949',
                                                                         ragged_rows=True,
                                                                         on_demand=True)
                                
                                # 시트 가져오기
                                sheet = workbook.sheet_by_name(self.sheet_name)
                                # 데이터 수동 추출
                                data = []
                                for row_idx in range(sheet.nrows):
                                    row_data = []
                                    for col_idx in range(sheet.ncols):
                                        cell_value = sheet.cell_value(row_idx, col_idx)
                                        row_data.append(cell_value)
                                    data.append(row_data)
                                # DataFrame 생성
                                self.df = pd.DataFrame(data)
                        else:
                            # 다른 옵션으로 시도
                            self.df = pd.read_excel(self.file_path, sheet_name=self.sheet_name, header=None, 
                                                  engine='openpyxl', na_filter=False)
                    except Exception as e2:
                        # 세 번째 방법: 파일 복사 후 시도 (손상된 파일 처리)
                        try:
                            # 임시 파일로 복사
                            temp_dir = tempfile.mkdtemp()
                            temp_file = os.path.join(temp_dir, os.path.basename(self.file_path))
                            shutil.copy2(self.file_path, temp_file)
                            
                            if file_ext == '.xls':
                                # Excel 97-2003 워크시트(.xls) 파일 처리를 위한 특별 설정
                                import xlrd
                                xlrd.mmap_mode = False
                                
                                try:
                                    # 첫 번째 시도: 기본 설정으로 열기
                                    workbook = xlrd.open_workbook(temp_file, formatting_info=False, 
                                                                 on_demand=True)
                                except Exception as xls_e1:
                                    try:
                                        # 두 번째 시도: 인코딩 지정 (한글 파일)
                                        workbook = xlrd.open_workbook(temp_file, formatting_info=False, 
                                                                     encoding_override='cp949',
                                                                     on_demand=True)
                                    except Exception as xls_e2:
                                        try:
                                            # 세 번째 시도: 레거시 모드 (Excel 97-2003 전용)
                                            workbook = xlrd.open_workbook(temp_file, formatting_info=False, 
                                                                         ragged_rows=True,
                                                                         on_demand=True)
                                        except Exception as xls_e3:
                                            # 네 번째 시도: 모든 옵션 조합
                                            workbook = xlrd.open_workbook(temp_file, formatting_info=False, 
                                                                         encoding_override='cp949',
                                                                         ragged_rows=True,
                                                                         on_demand=True)
                                
                                # 시트 가져오기
                                sheet = workbook.sheet_by_name(self.sheet_name)
                                # 데이터 수동 추출
                                data = []
                                for row_idx in range(sheet.nrows):
                                    row_data = []
                                    for col_idx in range(sheet.ncols):
                                        cell_value = sheet.cell_value(row_idx, col_idx)
                                        row_data.append(cell_value)
                                    data.append(row_data)
                                # DataFrame 생성
                                self.df = pd.DataFrame(data)
                            else:
                                # openpyxl 다양한 옵션 시도
                                from openpyxl import load_workbook
                                workbook = load_workbook(temp_file, data_only=True, read_only=True, keep_vba=False)
                                self.df = pd.read_excel(temp_file, sheet_name=self.sheet_name, header=None, 
                                                      engine='openpyxl')
                                
                            # 임시 디렉토리 정리
                            try:
                                shutil.rmtree(temp_dir)
                            except:
                                pass
                        except Exception as e3:
                            # 상세한 오류 정보 수집
                            import traceback
                            error_traceback = traceback.format_exc()
                            
                            detailed_error = {
                                "file_path": self.file_path,
                                "file_size": os.path.getsize(self.file_path) if os.path.exists(self.file_path) else "파일 없음",
                                "file_extension": file_ext,
                                "sheet_name": self.sheet_name,
                                "error1": f"기본 방식 오류: {str(e1)}",
                                "error2": f"다양한 옵션 시도 오류: {str(e2)}",
                                "error3": f"임시 파일 복사 오류: {str(e3)}",
                                "traceback": error_traceback
                            }
                            
                            # 상세 오류 메시지 생성
                            error_msg = f"시트 데이터 로딩 중 오류 발생:\n"
                            error_msg += f"파일: {self.file_path}\n"
                            error_msg += f"파일 크기: {detailed_error['file_size']} 바이트\n"
                            error_msg += f"파일 확장자: {detailed_error['file_extension']}\n"
                            error_msg += f"시트 이름: {detailed_error['sheet_name']}\n"
                            error_msg += f"오류 1: {detailed_error['error1']}\n"
                            error_msg += f"오류 2: {detailed_error['error2']}\n"
                            error_msg += f"오류 3: {detailed_error['error3']}\n"
                            error_msg += f"오류 위치: {detailed_error['traceback']}"
                            
                            # 오류 로그 파일에 기록 (디버깅용)
                            try:
                                log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
                                os.makedirs(log_dir, exist_ok=True)
                                log_file = os.path.join(log_dir, "excel_errors.log")
                                
                                with open(log_file, "a", encoding="utf-8") as f:
                                    f.write(f"\n{'='*50}\n")
                                    f.write(f"시간: {pd.Timestamp.now()}\n")
                                    f.write(error_msg)
                                    f.write(f"\n{'='*50}\n")
                                
                                # 로그 파일 경로 추가
                                error_msg += f"\n\n상세 로그가 다음 파일에 저장되었습니다: {log_file}"
                            except Exception as log_e:
                                error_msg += f"\n\n로그 파일 저장 실패: {str(log_e)}"
                            
                            QMessageBox.critical(self, '오류', error_msg)
                            return
                
                if self.df.empty:
                    return
                    
                # 첫 번째 행은 헤더로 사용
                self.header_row = self.df.iloc[0]
                
                # 테이블 업데이트
                self.update_table()
                
        except Exception as e:
            # 상세한 오류 정보 수집
            import traceback
            error_traceback = traceback.format_exc()
            
            # 파일 정보 수집
            file_info = {
                "file_path": self.file_path,
                "file_size": os.path.getsize(self.file_path) if os.path.exists(self.file_path) else "파일 없음",
                "file_extension": Path(self.file_path).suffix.lower(),
                "sheet_name": self.sheet_name,
                "error": str(e),
                "error_type": type(e).__name__,
                "traceback": error_traceback
            }
            
            # 상세 오류 메시지 생성
            error_msg = f"시트 데이터 로딩 중 오류 발생:\n"
            error_msg += f"파일: {file_info['file_path']}\n"
            error_msg += f"파일 크기: {file_info['file_size']} 바이트\n"
            error_msg += f"파일 확장자: {file_info['file_extension']}\n"
            error_msg += f"시트 이름: {file_info['sheet_name']}\n"
            error_msg += f"오류 유형: {file_info['error_type']}\n"
            error_msg += f"오류 내용: {file_info['error']}\n"
            error_msg += f"오류 위치: {file_info['traceback']}"
            
            # 오류 로그 파일에 기록 (디버깅용)
            try:
                log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
                os.makedirs(log_dir, exist_ok=True)
                log_file = os.path.join(log_dir, "excel_errors.log")
                
                with open(log_file, "a", encoding="utf-8") as f:
                    f.write(f"\n{'='*50}\n")
                    f.write(f"시간: {pd.Timestamp.now()}\n")
                    f.write(error_msg)
                    f.write(f"\n{'='*50}\n")
                
                # 로그 파일 경로 추가
                error_msg += f"\n\n상세 로그가 다음 파일에 저장되었습니다: {log_file}"
            except Exception as log_e:
                error_msg += f"\n\n로그 파일 저장 실패: {str(log_e)}"
            
            QMessageBox.critical(self, '오류', error_msg)
    
    def check_excel_process_windows(self, filename):
        """Windows에서 엑셀 프로세스 확인"""
        if not PSUTIL_AVAILABLE:
            return False
            
        try:
            filename_lower = filename.lower()
            for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
                try:
                    if proc.info['name'] and 'excel' in proc.info['name'].lower():
                        # 명령줄 인수에서 파일명 확인
                        if proc.info['cmdline']:
                            for arg in proc.info['cmdline']:
                                if filename_lower in arg.lower():
                                    return True
                        
                        # Windows API를 사용하여 창 제목 확인
                        if WIN32_AVAILABLE:
                            try:
                                def enum_windows_callback(hwnd, windows):
                                    if win32gui.IsWindowVisible(hwnd):
                                        window_title = win32gui.GetWindowText(hwnd)
                                        if window_title and filename_lower in window_title.lower():
                                            windows.append(hwnd)
                                    return True
                                
                                windows = []
                                win32gui.EnumWindows(enum_windows_callback, windows)
                                if windows:
                                    return True
                            except:
                                pass
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    continue
            return False
        except Exception:
            return False
    
    def check_excel_process_generic(self, filename):
        """일반적인 프로세스 확인 (macOS, Linux)"""
        if not PSUTIL_AVAILABLE:
            return False
            
        try:
            filename_lower = filename.lower()
            for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
                try:
                    if proc.info['name'] and ('excel' in proc.info['name'].lower() or 
                                            'libreoffice' in proc.info['name'].lower() or
                                            'calc' in proc.info['name'].lower()):
                        if proc.info['cmdline']:
                            for arg in proc.info['cmdline']:
                                if filename_lower in arg.lower():
                                    return True
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    continue
            return False
        except Exception:
            return False
    
    def check_excel_opened(self, file_path):
        """엑셀 파일이 열렸는지 확인"""
        filename = os.path.basename(file_path)
        
        if platform.system() == "Windows":
            return self.check_excel_process_windows(filename)
        else:
            return self.check_excel_process_generic(filename)
    
    def open_file_smart(self, file_path):
        """스마트 파일 열기 (버튼 상태 변경 방식)"""
        try:
            # 파일 존재 확인
            if not os.path.exists(file_path):
                QMessageBox.warning(self, '경고', f'파일을 찾을 수 없습니다:\n{file_path}')
                return
            
            # 버튼 상태 변경 (비활성화 및 텍스트 변경)
            original_text = self.sender().text() if self.sender() else "파일 열기"
            button = self.sender() if self.sender() else self.open_file_btn
            
            button.setEnabled(False)
            button.setText("여는 중...")
            
            # UI 업데이트 강제 실행
            from PyQt5.QtWidgets import QApplication
            QApplication.processEvents()
            
            # 운영체제별 파일 열기
            system = platform.system()
            
            if system == "Windows":
                os.startfile(file_path)
            elif system == "Darwin":
                subprocess.run(["open", file_path])
            else:
                subprocess.run(["xdg-open", file_path])
            
            # 설정된 시간 후 버튼 상태 복원
            QTimer.singleShot(FILE_OPEN_TIMEOUT, lambda: self.restore_button_state(button, original_text))
            
        except Exception as e:
            # 오류 발생 시 버튼 상태 즉시 복원
            if 'button' in locals() and 'original_text' in locals():
                self.restore_button_state(button, original_text)
            QMessageBox.critical(self, '오류', f'파일 열기 중 오류가 발생했습니다:\n{str(e)}')
    
    def restore_button_state(self, button, original_text):
        """버튼 상태 복원"""
        try:
            button.setEnabled(True)
            button.setText(original_text)
        except Exception:
            # 버튼이 이미 삭제된 경우 등의 예외 상황 무시
            pass
    
    def monitor_excel_process(self, loading_msg, file_path):
        """엑셀 프로세스 모니터링"""
        self.check_count = 0
        self.max_checks = EXCEL_PROCESS_MAX_CHECKS
        self.loading_msg = loading_msg
        self.monitoring_file_path = file_path
        
        # 설정된 간격 후 첫 번째 체크 시작
        QTimer.singleShot(EXCEL_PROCESS_CHECK_INTERVAL, self.check_excel_process_timer)
    
    def check_excel_process_timer(self):
        """타이머 기반 엑셀 프로세스 체크"""
        try:
            self.check_count += 1
            
            # 엑셀이 열렸는지 확인
            if self.check_excel_opened(self.monitoring_file_path):
                # 엑셀이 열렸으면 로딩 메시지 닫기
                if hasattr(self, 'loading_msg') and self.loading_msg:
                    self.loading_msg.close()
                return
            
            # 최대 체크 횟수에 도달했으면 강제 닫기
            if self.check_count >= self.max_checks:
                if hasattr(self, 'loading_msg') and self.loading_msg:
                    self.loading_msg.close()
                return
            
            # 설정된 간격 후 다시 체크
            QTimer.singleShot(EXCEL_PROCESS_CHECK_INTERVAL, self.check_excel_process_timer)
            
        except Exception:
            # 오류 발생 시 로딩 메시지 닫기
            if hasattr(self, 'loading_msg') and self.loading_msg:
                self.loading_msg.close()
    
    def open_file(self):
        """파일을 기본 프로그램으로 열기"""
        self.open_file_smart(self.file_path)
    
    def open_file_writable(self):
        """파일의 읽기 전용 속성을 해제한 후 열기"""
        try:
            # 파일 존재 확인
            if not os.path.exists(self.file_path):
                QMessageBox.warning(self, '경고', f'파일을 찾을 수 없습니다:\n{self.file_path}')
                return
            
            # 현재 파일 권한 확인
            current_permissions = os.stat(self.file_path).st_mode
            
            # 읽기 전용인지 확인 (쓰기 권한이 없는지 확인)
            is_readonly = not (current_permissions & stat.S_IWRITE)
            
            if is_readonly:
                try:
                    # 읽기 전용 속성 해제 (쓰기 권한 추가)
                    new_permissions = current_permissions | stat.S_IWRITE
                    os.chmod(self.file_path, new_permissions)
                    
                except PermissionError:
                    QMessageBox.warning(self, '경고', 
                                      f'파일의 읽기 전용 속성을 해제할 권한이 없습니다:\n{os.path.basename(self.file_path)}\n\n관리자 권한으로 실행하거나 파일 소유자에게 문의하세요.')
                    return
                except Exception as e:
                    QMessageBox.critical(self, '오류', 
                                       f'파일 권한 변경 중 오류가 발생했습니다:\n{str(e)}')
                    return
            
            # 스마트 파일 열기 사용
            self.open_file_smart(self.file_path)
            
        except Exception as e:
            QMessageBox.critical(self, '오류', f'파일 처리 중 오류가 발생했습니다:\n{str(e)}')
    
    def open_folder(self):
        """파일이 있는 폴더를 탐색기로 열기"""
        try:
            # 파일 존재 확인
            if not os.path.exists(self.file_path):
                QMessageBox.warning(self, '경고', f'파일을 찾을 수 없습니다:\n{self.file_path}')
                return
            
            # 폴더 경로 추출
            folder_path = os.path.dirname(self.file_path)
            
            # 운영체제별 폴더 열기
            system = platform.system()
            
            if system == "Windows":
                # Windows: explorer /select 사용하여 파일이 선택된 상태로 폴더 열기
                subprocess.run(["explorer", "/select,", self.file_path])
            elif system == "Darwin":  # macOS
                # macOS: Finder에서 파일 선택
                subprocess.run(["open", "-R", self.file_path])
            else:  # Linux 및 기타
                # Linux: 파일 매니저로 폴더 열기 (파일 선택은 지원하지 않을 수 있음)
                try:
                    # 먼저 nautilus 시도 (Ubuntu/GNOME)
                    subprocess.run(["nautilus", "--select", self.file_path])
                except (subprocess.CalledProcessError, FileNotFoundError):
                    try:
                        # dolphin 시도 (KDE)
                        subprocess.run(["dolphin", "--select", self.file_path])
                    except (subprocess.CalledProcessError, FileNotFoundError):
                        # 기본 파일 매니저로 폴더만 열기
                        subprocess.run(["xdg-open", folder_path])
            
        except Exception as e:
            QMessageBox.critical(self, '오류', f'폴더 열기 중 오류가 발생했습니다:\n{str(e)}')
